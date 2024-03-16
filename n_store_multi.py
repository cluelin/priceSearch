import requests
import openpyxl
import logging
import json
import threading
import time

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

header = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
          'X-Naver-Client-Id' : 'XiloPbctZ5cNmCtyUSjN',
          'X-Naver-Client-Secret' : '5mmboFr6ji'}

search_term_index = 1
result_term_index = 2
result_price_index = 3
result_url_index = 4

def search_on_engine(term_list:list, search_result_list, start_index, end_index):    
    for i in range(start_index, end_index):
        logging.debug(f"검색어: {term_list[i]}")

        if term_list[i]:
            search_url = f"https://openapi.naver.com/v1/search/shop.json?query={term_list[i]}&exclude=used:rental:cbshop&sort=asc&display=1" 
            response = requests.get(search_url, headers=header)
            
            if response.status_code == 200:
                search_result = json.loads(response.text)

                if search_result.get('items'):
                    title = search_result.get('items')[0].get('title')
                    price = search_result.get('items')[0].get('lprice')
                    link = search_result.get('items')[0].get('link')

                    logging.debug(f"title : {title}, price : {price}, link = {link}")

                    title = title.replace('<b>', '')
                    title = title.replace('</b>', '')

                    search_result_list[i] = {'title' : title, 'price' : price, 'link' : link }
                else:
                    search_result_list[i] = None
            else:
                logging.error(f"url : {search_url}, response : {response}")
                search_result_list[i] = None


def read(filename):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    search_term_list = []
    # 데이터 검색 및 결과 엑셀에 저장
    for row in range(2, sheet.max_row + 1):
        search_term = sheet.cell(row, search_term_index).value  # 엑셀에서 검색어 가져오기
        search_term_list.append(search_term)

    return search_term_list, wb

def save(wb, filename, price_matrix):
    
    sheet = wb.active

    for index, item in enumerate(price_matrix):
        if item:
            # 검색 결과에서 필요한 정보를 추출하고 검색 결과 엑셀에 저장
            sheet.cell(index+2 , result_term_index, item.get('title'))
            sheet.cell(index+2, result_price_index, item.get('price'))
            cell = sheet.cell(index+2, result_url_index, item.get('link'))
            cell.hyperlink = item.get('link')
            cell.style = "Hyperlink"

    # 변경 내용을 저장하고 파일 닫기
    wb.save(filename)
    wb.close()


def search_multi(search_term_list):
    length = len(search_term_list)
    search_result_list = list(range(length))
        
    threads = []
    thread_count = 5
    chunk_size = length // thread_count

    for i in range(thread_count):
        start_index = i * chunk_size
        end_index = (i+1) * chunk_size
        if i == 4:
            end_index = length
        thread = threading.Thread(target=search_on_engine, 
                                  args=(search_term_list, search_result_list, start_index , end_index))
        threads.append(thread)
        thread.start()

    # 모든 스레드가 종료될 때까지 대기
    for thread in threads:
        thread.join()

    return search_result_list


# 메인 함수
def main():
    # 검색을 실행할 엑셀 파일과 쿼리를 설정
    filename = "data.xlsx"

    start_time = time.time()

    # 검색 및 저장 함수 호출
    search_term_list, wb = read(filename)
    read_time = time.time()
    logging.info(f"read 종료 시간 : {read_time - start_time}")


    search_result_list = search_multi(search_term_list)
    search_time = time.time()
    logging.info(f"search 종료 시간 : {search_time - read_time}")

    save(wb, filename, search_result_list)
    save_time = time.time()
    logging.info(f"save 종료 시간 : {save_time - search_time}")

if __name__ == "__main__":
    main()
