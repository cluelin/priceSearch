import requests
from bs4 import BeautifulSoup
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

header = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0'}

def search_and_save(query, filename):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # 결과를 저장할 새로운 컬럼 추가
    sheet.cell(1, sheet.max_column + 1, "검색 결과")

    # 데이터 검색 및 결과 엑셀에 저장
    for row in range(2, sheet.max_row + 1):
        search_term = sheet.cell(row, 1).value  # 엑셀에서 검색어 가져오기

        logging.info(f"검색어: {search_term}")

        # 검색어가 있을 경우에만 검색을 실행
        if search_term:
            search_url = f"https://search.danawa.com/dsearch.php?k1={search_term}&module=goods&act=dispMain"  # 예시 검색 엔진 URL
            response = requests.get(search_url, headers=header, verify=False)
            
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                # 검색 결과에서 필요한 정보를 추출하고 검색 결과 엑셀에 저장
                search_result = soup.find("div", class_="search-result").text  # 예시: 검색 결과가 div 태그의 search-result 클래스에 있음
                sheet.cell(row, sheet.max_column, search_result)
            else:
                logging.info(f"url : {search_url}, response : {response}")
                sheet.cell(row, sheet.max_column, "검색 실패")

    # 변경 내용을 저장하고 파일 닫기
    wb.save(filename)
    wb.close()

# 메인 함수
def main():
    # 검색을 실행할 엑셀 파일과 쿼리를 설정
    excel_file = "data.xlsx"
    query = "상품명"  # 예시: 엑셀의 첫 번째 열에 검색어가 있는 경우

    # 검색 및 저장 함수 호출
    search_and_save(query, excel_file)

if __name__ == "__main__":
    main()
